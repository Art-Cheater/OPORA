"""Импорт заявок из Excel «Заявки по энергосервису»."""

from __future__ import annotations

import re
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from app.core.audit_service import AuditService
from app.core.exceptions import ValidationError
from app.extensions import db
from app.models.base import utcnow
from app.models.enums import AuditAction, EntityType, Priority
from app.models.requests.request import Request
from app.modules.requests.address_format import format_address, normalize_address
from app.modules.requests.services import RequestService
from app.modules.requests.workflow import STATUS_NEW

_DUE_RE = re.compile(
    r"срок\s+исполнения\s+до\s+(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,4})",
    re.IGNORECASE,
)
_HEADER_ALIASES = {
    "street": ("улица", "ул"),
    "yard": ("двор",),
    "count": ("кол-во", "колво", "количество", "звон"),
    "journal": ("заявк", "номер"),
    "pp": ("пп", "пункт питания"),
}
_DONE_RE = re.compile(r"сделано", re.IGNORECASE)
_BARRIER_RE = re.compile(r"шлагбаум", re.IGNORECASE)
_WATT_RE = re.compile(r"\d+\s*вт|днат", re.IGNORECASE)
_NOT_ES_RE = re.compile(r"не\s+энергосервис", re.IGNORECASE)


@dataclass
class EnergoserviceRow:
    raw_address: str
    kind: str  # street | yard
    call_count: int
    journal_numbers: str | None
    pp: str | None
    due_date: date | None
    notes: list[str] = field(default_factory=list)
    has_barrier: bool = False
    is_done: bool = False
    not_energoservice: bool = False
    sheet: str = ""
    excel_row: int = 0


@dataclass
class EnergoserviceImportResult:
    created: int = 0
    skipped: int = 0
    total: int = 0


def _cell_text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    if text.endswith(".0") and text.replace(".", "", 1).isdigit():
        text = text[:-2]
    return text


def _norm_header(value) -> str:
    return re.sub(r"\s+", " ", _cell_text(value).casefold())


def _as_count(value) -> int:
    text = _cell_text(value).replace(",", ".")
    if not text:
        return 1
    try:
        number = float(text)
    except ValueError:
        digits = re.search(r"\d+", text)
        return max(1, int(digits.group())) if digits else 1
    return max(1, int(number))


def _parse_due(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = _cell_text(value)
    if not text:
        return None
    match = _DUE_RE.search(text)
    if not match:
        return None
    day, month, year = (int(match.group(1)), int(match.group(2)), int(match.group(3)))
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _header_map(values: tuple) -> dict[str, int]:
    found: dict[str, int] = {}
    for idx, raw in enumerate(values or ()):
        header = _norm_header(raw)
        if not header:
            continue
        for key, aliases in _HEADER_ALIASES.items():
            if key in found:
                continue
            if any(alias in header for alias in aliases):
                found[key] = idx
                break
    return found


def _row_notes(*values) -> list[str]:
    notes: list[str] = []
    for value in values:
        text = _cell_text(value)
        if not text:
            continue
        if _DUE_RE.search(text):
            continue
        notes.append(text)
    return notes


def parse_energoservice_xlsx(path: Path) -> list[EnergoserviceRow]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ValidationError("Для импорта нужен пакет openpyxl.") from exc

    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[EnergoserviceRow] = []
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        header_idx = None
        cols: dict[str, int] = {}
        for idx, values in enumerate(worksheet.iter_rows(max_row=15, values_only=True), start=1):
            mapped = _header_map(values or ())
            if "street" in mapped or "yard" in mapped:
                header_idx = idx
                cols = mapped
                break
        if header_idx is None:
            continue

        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=header_idx + 1, values_only=True),
            start=header_idx + 1,
        ):
            values = values or ()
            street = _cell_text(values[cols["street"]]) if "street" in cols and cols["street"] < len(values) else ""
            yard = _cell_text(values[cols["yard"]]) if "yard" in cols and cols["yard"] < len(values) else ""
            if street and yard:
                raw_address = f"{street} (двор: {yard})"
                kind = "street"
            elif street:
                raw_address = street
                kind = "street"
            elif yard:
                raw_address = yard
                kind = "yard"
            else:
                continue

            count_val = values[cols["count"]] if "count" in cols and cols["count"] < len(values) else None
            journal = _cell_text(values[cols["journal"]]) if "journal" in cols and cols["journal"] < len(values) else ""
            pp = _cell_text(values[cols["pp"]]) if "pp" in cols and cols["pp"] < len(values) else ""

            due = None
            extra_cells = []
            mapped_idxs = set(cols.values())
            for idx, cell in enumerate(values):
                if idx in mapped_idxs:
                    continue
                extra_cells.append(cell)
                due = due or _parse_due(cell)

            notes = _row_notes(*extra_cells)
            blob = " ".join(notes)
            rows.append(
                EnergoserviceRow(
                    raw_address=raw_address,
                    kind=kind,
                    call_count=_as_count(count_val),
                    journal_numbers=journal or None,
                    pp=pp or None,
                    due_date=due,
                    notes=notes,
                    has_barrier=bool(_BARRIER_RE.search(blob)),
                    is_done=bool(_DONE_RE.search(blob)),
                    not_energoservice=bool(_NOT_ES_RE.search(blob)),
                    sheet=sheet_name,
                    excel_row=excel_row,
                )
            )
    workbook.close()
    return rows


def _description(row: EnergoserviceRow) -> str:
    lines = [
        "Импорт из Excel «Заявки по энергосервису». Карточка неполная — дозаполнить вручную.",
        f"Тип адреса: {'двор' if row.kind == 'yard' else 'улица'}.",
        f"Звонков по журналу: {row.call_count}.",
    ]
    if row.journal_numbers:
        lines.append(f"№ в журнале: {row.journal_numbers}.")
    if row.due_date:
        lines.append(f"Срок исполнения: {row.due_date.strftime('%d.%m.%Y')}.")
    if row.is_done:
        lines.append("В файле отмечено: сделано.")
    if row.not_energoservice:
        lines.append("В файле отмечено: не энергосервисный.")
    extra = [note for note in row.notes if not _DONE_RE.fullmatch(note)]
    if extra:
        lines.append("Пометки: " + "; ".join(extra) + ".")
    return "\n".join(lines)


class EnergoserviceImportService:
    @classmethod
    def import_from_xlsx(cls, path: Path, actor_id: uuid.UUID) -> EnergoserviceImportResult:
        parsed = parse_energoservice_xlsx(path)
        result = EnergoserviceImportResult(total=len(parsed))
        if not parsed:
            raise ValidationError("В файле нет строк с адресом (колонки «Улица» / «Двор»).")

        status = RequestService.get_status_by_code(STATUS_NEW)
        existing = {
            (item or "").strip()
            for item in db.session.scalars(
                db.select(Request.normalized_address).where(
                    Request.active_filter(),
                    Request.normalized_address.is_not(None),
                )
            )
            if item
        }
        year = datetime.now().year
        seq = cls._next_seq(year)

        created: list[Request] = []
        for row in parsed:
            formatted = format_address(row.raw_address)
            key = normalize_address(formatted)
            if not key or key in existing:
                result.skipped += 1
                continue
            seq += 1
            number = f"REQ-{year}-{seq:03d}" if seq < 1000 else f"REQ-{year}-{seq}"
            received = utcnow()
            req = Request(
                number=number,
                title=formatted[:500],
                description=_description(row),
                address=formatted[:500],
                original_address=row.raw_address[:500],
                normalized_address=key[:1000],
                pp=(row.pp or "")[:255] or None,
                received_at=received,
                dispatcher_name=None,
                applicant_name="—",
                has_barrier=row.has_barrier,
                barrier_phone="не указан" if row.has_barrier else None,
                repeat_count=row.call_count,
                repeat_dates=[],
                priority=Priority.MEDIUM.value,
                status_id=status.id,
                due_date=row.due_date,
                address_source="energoservice_xlsx",
                created_by=actor_id,
                updated_by=actor_id,
            )
            db.session.add(req)
            existing.add(key)
            created.append(req)
            result.created += 1

        if created:
            db.session.flush()
            AuditService.log(
                user_id=actor_id,
                action=AuditAction.CREATE.value,
                entity_type=EntityType.REQUEST.value,
                entity_id=created[0].id,
                description=(
                    f"Импорт энергосервиса: создано {result.created}, "
                    f"пропущено {result.skipped}, строк {result.total}"
                ),
            )
        db.session.commit()
        return result

    @staticmethod
    def _next_seq(year: int) -> int:
        prefix = f"REQ-{year}-"
        last = db.session.scalar(
            db.select(Request.number)
            .where(Request.number.ilike(f"{prefix}%"))
            .order_by(Request.number.desc())
            .limit(1)
        )
        if not last:
            return 0
        try:
            return int(last.rsplit("-", 1)[-1])
        except ValueError:
            return 0
