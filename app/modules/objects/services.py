"""Сервисы модуля объектов."""

from __future__ import annotations

import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from app.core.audit_service import AuditService
from app.core.exceptions import ValidationError
from app.extensions import db
from app.models.enums import (
    AuditAction,
    EntityType,
    ProjectStatus,
    WorkObjectKind,
    WorkObjectStatus,
)
from app.models.projects.project import Project
from app.models.work_objects.work_object import WorkObject
from app.modules.projects.repositories import ProjectRepository
from app.modules.projects.services import ProjectPayload, ProjectService

WORK_TYPE_DEFAULT = "Устройство наружного освещения"
AUTO_PROJECT_RESULT = (
    "Обследование проведено, ТЗ подготовлено, локально-сметный расчет готов."
)
TZ_RESULT_ALT = "Подготовлено техническое задание и локально-сметный расчет"
_CLOSED_PROJECT_STATUSES = (
    ProjectStatus.COMPLETED.value,
    ProjectStatus.CANCELLED.value,
    ProjectStatus.ARCHIVED.value,
)

# Результат из плана → предлагаемый статус проекта
_RESULT_DRAFT_RE = re.compile(
    r"обследование\s+проведено.*тз\s+подготовлено|локально.?сметн",
    re.IGNORECASE,
)
_RESULT_ACTIVE_RE = re.compile(
    r"подготовка\s+рабочей\s+документации|ид[её]т\s+подготовка\s+рабочей",
    re.IGNORECASE,
)
_RESULT_TENDER_RE = re.compile(
    r"в\s+закупках|заявка\s+у\s+экономистов",
    re.IGNORECASE,
)
_RESULT_CLOSED_RE = re.compile(r"принят|выполнен", re.IGNORECASE)

# Строки-мусор внизу листов Excel (подписи, «План/Остаток» и т.п.)
_JUNK_NAME_RE = re.compile(
    r"^(план|остаток|куратор|начальник|и\.?\s*о\.?|туров|телефон|\d[\d\-\s]{5,})$",
    re.IGNORECASE,
)


@dataclass
class ObjectPayload:
    name: str
    work_type: str | None = None
    object_kind: str | None = None
    address: str | None = None
    plan_year: int | None = None
    work_deadline: str | None = None
    contract_number: str | None = None
    contract_date: date | None = None
    contractor_name: str | None = None
    contract_amount: Decimal | None = None
    budget_amount: Decimal | None = None
    court_decision_number: str | None = None
    kind_comment: str | None = None
    result_text: str | None = None
    source_sheet: str | None = None
    notes: str | None = None
    status: str = WorkObjectStatus.FREE.value
    create_draft_project: bool = False


@dataclass
class ImportResult:
    created: int = 0
    updated: int = 0
    skipped: int = 0
    total: int = 0


class ObjectService:
    @staticmethod
    def _normalized_result(value: str | None) -> str:
        """Нормализовать пробелы для точного сопоставления результата."""
        return re.sub(r"\s+", " ", value or "").strip().casefold()

    @classmethod
    def should_create_project(cls, result_text: str | None) -> bool:
        return cls._needs_project(result_text)

    @classmethod
    def _needs_project(cls, result_text: str | None) -> bool:
        if cls._is_closed_result(result_text):
            return False
        folded = cls._normalized_result(result_text)
        if not folded:
            return False
        if folded == cls._normalized_result(AUTO_PROJECT_RESULT):
            return True
        if folded == cls._normalized_result(TZ_RESULT_ALT):
            return True
        return bool(_RESULT_DRAFT_RE.search(result_text or ""))

    @classmethod
    def _needs_tender(cls, result_text: str | None) -> bool:
        if cls._is_closed_result(result_text):
            return False
        return bool(_RESULT_TENDER_RE.search(result_text or ""))

    @classmethod
    def _is_closed_result(cls, result_text: str | None) -> bool:
        return bool(_RESULT_CLOSED_RE.search(result_text or ""))

    @classmethod
    def _needs_contract(cls, obj: WorkObject) -> bool:
        if not (obj.contract_number or "").strip():
            return False
        return not cls._is_closed_result(obj.result_text)

    @classmethod
    def _active_project(cls, obj: WorkObject) -> Project | None:
        return db.session.scalar(
            db.select(Project)
            .where(
                Project.object_id == obj.id,
                Project.active_filter(),
                Project.status.notin_(_CLOSED_PROJECT_STATUSES),
            )
            .order_by(Project.created_at.asc())
            .limit(1)
        )

    @classmethod
    def _active_tender(cls, obj: WorkObject, project: Project | None = None):
        from app.models.tenders.tender_application import TenderApplication
        from app.models.tenders.tender_project import TenderProject

        tender = db.session.scalar(
            db.select(TenderApplication)
            .where(
                TenderApplication.object_id == obj.id,
                TenderApplication.active_filter(),
            )
            .order_by(TenderApplication.created_at.asc())
            .limit(1)
        )
        if tender is not None:
            return tender
        if project is None:
            return None
        return db.session.scalar(
            db.select(TenderApplication)
            .join(TenderProject, TenderProject.tender_id == TenderApplication.id)
            .where(
                TenderProject.project_id == project.id,
                TenderProject.active_filter(),
                TenderApplication.active_filter(),
            )
            .order_by(TenderApplication.created_at.asc())
            .limit(1)
        )

    @classmethod
    def _active_contract(cls, obj: WorkObject):
        from sqlalchemy import case, desc, nulls_last

        from app.models.contracts.contract import Contract
        from app.models.contracts.contract_object import ContractObject
        from app.models.enums import ContractStatus

        active_statuses = (
            ContractStatus.ACTIVE.value,
            ContractStatus.WORK_DOCS_PENDING.value,
            ContractStatus.IN_PROGRESS.value,
            ContractStatus.KS2_PENDING.value,
            ContractStatus.REJECTED.value,
        )
        priority = case(
            (Contract.status.in_(active_statuses), 0),
            (Contract.status == ContractStatus.COMPLETED.value, 1),
            else_=2,
        )
        return db.session.scalar(
            db.select(Contract)
            .join(ContractObject, ContractObject.contract_id == Contract.id)
            .where(
                ContractObject.object_id == obj.id,
                Contract.active_filter(),
            )
            .order_by(
                priority,
                nulls_last(desc(Contract.contract_date)),
                desc(Contract.created_at),
            )
            .limit(1)
        )

    @classmethod
    def related_projects(cls, obj: WorkObject) -> list:
        from app.models.projects.project import Project

        return list(
            db.session.scalars(
                db.select(Project)
                .where(
                    Project.object_id == obj.id,
                    Project.active_filter(),
                )
                .order_by(Project.created_at.desc())
            )
        )

    @classmethod
    def related_contracts(cls, obj: WorkObject) -> list:
        """Все контракты, связанные с объектом (новые сверху)."""
        from sqlalchemy import desc, nulls_last

        from app.models.contracts.contract import Contract
        from app.models.contracts.contract_object import ContractObject

        return list(
            db.session.scalars(
                db.select(Contract)
                .join(ContractObject, ContractObject.contract_id == Contract.id)
                .where(
                    ContractObject.object_id == obj.id,
                    ContractObject.active_filter(),
                    Contract.active_filter(),
                )
                .order_by(
                    nulls_last(desc(Contract.contract_date)),
                    desc(Contract.created_at),
                )
            )
        )

    @classmethod
    def related_chain(cls, obj: WorkObject) -> dict:
        """Активные проект / заявка / контракт объекта — для ссылок с карточки."""
        project = cls._active_project(obj)
        contracts = cls.related_contracts(obj)
        return {
            "project": project,
            "projects": cls.related_projects(obj),
            "tender": cls._active_tender(obj, project),
            "contract": cls._active_contract(obj),
            "contracts": contracts,
        }

    @classmethod
    def _apply_plan_volumes(cls, project: Project, volumes: dict | None) -> None:
        if not volumes:
            return
        if volumes.get("sip_meters") is not None:
            project.sip_meters = volumes["sip_meters"]
        if volumes.get("poles_count") is not None:
            project.poles_count = volumes["poles_count"]
        if volumes.get("lights_count") is not None:
            project.lights_count = volumes["lights_count"]
        if volumes.get("shuno_count") is not None:
            project.shuno_count = volumes["shuno_count"]

    @classmethod
    def _ensure_project_for_result(
        cls,
        obj: WorkObject,
        user_id: uuid.UUID,
        volumes: dict | None = None,
    ) -> Project | None:
        """Совместимость: создать проект по результату ТЗ/ЛСР."""
        return cls._ensure_chain_for_result(obj, user_id, volumes=volumes)

    @classmethod
    def _ensure_draft_project(cls, obj: WorkObject, user_id: uuid.UUID) -> Project | None:
        """Создать один черновик проекта, если у объекта его ещё нет."""
        existing = cls._active_project(obj)
        if existing is not None:
            return existing
        db.session.flush()
        project = ProjectService.create_project(
            ProjectPayload(
                code=ProjectRepository.next_code(),
                name=(obj.display_address or obj.name)[:500],
                description=obj.result_text or obj.kind_comment,
                status=ProjectStatus.DRAFT.value,
                progress_percent=0,
                start_date=None,
                end_date=None,
                responsible_id=user_id,
                executor_ids=[],
                object_id=obj.id,
            ),
            user_id,
            commit=False,
            allow_busy_object=True,
        )
        obj.status = WorkObjectStatus.IN_PROJECT.value
        obj.updated_by = user_id
        return project

    @classmethod
    def _ensure_chain_for_result(
        cls,
        obj: WorkObject,
        user_id: uuid.UUID,
        volumes: dict | None = None,
    ) -> Project | None:
        """
        Идемпотентно создать черновики по колонке «Результат» и номеру контракта.

        Метод не фиксирует транзакцию.
        """
        if obj.status == WorkObjectStatus.COMPLETED.value and cls._is_closed_result(obj.result_text):
            return cls._active_project(obj)

        need_contract = cls._needs_contract(obj)
        need_tender = need_contract or cls._needs_tender(obj.result_text)
        need_project = need_tender or need_contract or cls._needs_project(obj.result_text)
        if not need_project:
            return None

        db.session.flush()
        locked_obj = db.session.scalar(
            db.select(WorkObject)
            .where(WorkObject.id == obj.id, WorkObject.active_filter())
            .with_for_update()
        )
        if locked_obj is not None:
            obj = locked_obj

        project = cls._active_project(obj)
        if project is None:
            project = ProjectService.create_project(
                ProjectPayload(
                    code=ProjectRepository.next_code(),
                    name=(obj.display_address or obj.name)[:500],
                    description=obj.result_text,
                    status=ProjectStatus.DRAFT.value,
                    progress_percent=0,
                    start_date=None,
                    end_date=None,
                    responsible_id=user_id,
                    executor_ids=[],
                    object_id=obj.id,
                    sip_meters=(volumes or {}).get("sip_meters"),
                    poles_count=(volumes or {}).get("poles_count"),
                    lights_count=(volumes or {}).get("lights_count"),
                    shuno_count=(volumes or {}).get("shuno_count"),
                ),
                user_id,
                commit=False,
                allow_busy_object=True,
            )
        else:
            cls._apply_plan_volumes(project, volumes)

        obj.status = WorkObjectStatus.IN_PROJECT.value
        obj.updated_by = user_id

        if not need_tender:
            return project

        from app.models.enums import TenderApplicationStatus
        from app.modules.tenders.repositories import TenderRepository
        from app.modules.tenders.services import TenderPayload, TenderService

        tender = cls._active_tender(obj, project)
        if tender is None:
            tender = TenderService.create(
                TenderPayload(
                    number=TenderRepository.next_number(),
                    title=(obj.display_address or obj.name)[:500],
                    description=obj.result_text,
                    status=TenderApplicationStatus.DRAFT.value,
                    responsible_id=user_id,
                    project_ids=[project.id],
                    object_id=obj.id,
                    work_deadline=obj.work_deadline,
                    published_at=None,
                ),
                user_id,
                commit=False,
            )
        obj.status = WorkObjectStatus.IN_TENDER.value
        obj.updated_by = user_id

        if not need_contract:
            return project

        from app.modules.contracts.services import ContractService

        ContractService.create_draft_from_plan(
            obj,
            user_id,
            project=project,
            tender=tender,
            commit=False,
        )
        return project

    @staticmethod
    def suggested_project_status(result_text: str | None) -> str | None:
        """
        По колонке «Результат» из плана:
        - ТЗ/сметный расчёт готов → черновик проекта
        - идёт подготовка рабочей документации → проект «В работе»
        """
        text = (result_text or "").strip()
        if not text:
            return None
        if _RESULT_ACTIVE_RE.search(text):
            return ProjectStatus.ACTIVE.value
        if _RESULT_DRAFT_RE.search(text) or _RESULT_TENDER_RE.search(text):
            return ProjectStatus.DRAFT.value
        return None

    @staticmethod
    def can_create_contract_from_plan(obj: WorkObject) -> bool:
        return bool((obj.contract_number or "").strip())

    @staticmethod
    def suggested_contract_amount(obj: WorkObject) -> Decimal | None:
        """
        Для формы контракта: сумма контракта, если есть;
        иначе НМЦК (бюджет). Это разные поля — при появлении контракта
        сумма контракта не подменяется на НМЦК.
        """
        if obj.contract_amount is not None:
            return obj.contract_amount
        return obj.budget_amount

    @staticmethod
    def detect_object_kind(sheet_name: str | None) -> str:
        """По названию листа Excel: плановый / судебный / тех. присоединение."""
        s = (sheet_name or "").casefold().replace("ё", "е").strip()
        if "судеб" in s:
            return WorkObjectKind.COURT.value
        if "тех" in s and ("прис" in s or "присоед" in s):
            return WorkObjectKind.TECH_CONNECT.value
        return WorkObjectKind.PLANNED.value

    @staticmethod
    def _normalize(value: str | None) -> str | None:
        if value is None:
            return None
        stripped = value.strip()
        return stripped if stripped else None

    @staticmethod
    def _as_text(value) -> str | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = re.sub(r"\s+", " ", str(value)).strip()
        return text or None

    @staticmethod
    def _as_date(value) -> date | None:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        text = str(value).strip()
        if not text:
            return None
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d.%m.%y"):
            try:
                return datetime.strptime(text[:10], fmt).date()
            except ValueError:
                continue
        m = re.search(r"(\d{2}\.\d{2}\.\d{4})", text)
        if m:
            try:
                return datetime.strptime(m.group(1), "%d.%m.%Y").date()
            except ValueError:
                return None
        return None

    @staticmethod
    def _as_decimal(value) -> Decimal | None:
        if value is None or value == "":
            return None
        if isinstance(value, Decimal):
            return value
        if isinstance(value, (int, float)):
            return Decimal(str(value))
        text = str(value).replace(" ", "").replace(",", ".")
        text = re.sub(r"[^\d.\-]", "", text)
        if not text:
            return None
        try:
            return Decimal(text)
        except InvalidOperation:
            return None

    @staticmethod
    def _as_int(value) -> int | None:
        if value is None or value == "":
            return None
        if isinstance(value, bool):
            return None
        if isinstance(value, int):
            return value
        decimal_value = ObjectService._as_decimal(value)
        if decimal_value is None:
            return None
        return int(decimal_value)

    @classmethod
    def split_type_and_address(cls, raw_name: str) -> tuple[str, str]:
        """
        «Устройство наружного освещения по ул. …» →
        тип = Устройство наружного освещения, адрес = ул. …
        """
        name = re.sub(r"\s+", " ", (raw_name or "").strip())
        patterns = [
            (
                r"устройство\s+наружного\s+освещени[яе]?",
                "Устройство наружного освещения",
            ),
            (
                r"устройство\s+недостающего\s+электрического\s+освещения",
                "Устройство недостающего электрического освещения",
            ),
            (
                r"устройство\s+недостающего\s+наружного\s+освещени[яе]?",
                "Устройство недостающего наружного освещения",
            ),
        ]
        for pat, canonical in patterns:
            m = re.match(rf"^({pat})\s*(?:по|в|на)?\s*(.*)$", name, flags=re.IGNORECASE)
            if not m:
                continue
            address = (m.group(2) or "").strip(" .,;")
            if not address:
                address = name
            if address.casefold().startswith("устройство"):
                cut = re.search(
                    r"\b(ул\.|улица|д\.|дер\.|п\.|пос\.|посёлок|поселок|сл\.|слобода|мкр\.|пр\.|проезд)\b.+",
                    name,
                    flags=re.IGNORECASE,
                )
                if cut:
                    address = cut.group(0).strip(" .,;")
            return canonical, address or name

        # Общий случай: «Устройство … по/в/на <адрес>»
        m2 = re.match(
            r"^(устройство\s+.+?)\s+(?:по|в|на)\s+(.+)$",
            name,
            flags=re.IGNORECASE,
        )
        if m2:
            return m2.group(1).strip(), m2.group(2).strip(" .,;")

        return WORK_TYPE_DEFAULT, name

    @staticmethod
    def _header_map(header_row: tuple) -> dict[str, int]:
        """Сопоставить колонки по ключевым словам в заголовке."""
        mapping: dict[str, int] = {}
        for idx, cell in enumerate(header_row or ()):
            if not isinstance(cell, str):
                continue
            h = cell.casefold().replace("ё", "е")
            if "наименование" in h:
                mapping["name"] = idx
            elif "срок" in h and "выполн" in h:
                mapping["deadline"] = idx
            elif "подрядчик" in h:
                mapping["contractor"] = idx
            elif "номер" in h and "контракт" in h and "дата" in h:
                mapping["contract_combo"] = idx
            elif "номер" in h and "контракт" in h:
                mapping["contract_number"] = idx
            elif ("заключен" in h or "заключение" in h) and "контракт" in h:
                mapping["contract_date"] = idx
            elif "сумма" in h and "контракт" in h:
                mapping["contract_amount"] = idx
            elif "судебн" in h:
                mapping["court_decision"] = idx
            elif "расход" in h or "нмцк" in h or "бюджет" in h:
                mapping["budget"] = idx
            elif "результат" in h:
                mapping["result"] = idx
            elif "примечан" in h:
                mapping["notes"] = idx
            elif "светильник" in h:
                mapping["lights"] = idx
            elif "опор" in h:
                mapping["poles"] = idx
            elif "сип" in h:
                mapping["sip"] = idx
            elif "шуно" in h or "шкаф" in h:
                mapping["shuno"] = idx
            elif h.strip().startswith("№") or "п/п" in h:
                mapping["num"] = idx
        return mapping

    @staticmethod
    def _is_data_row(row: tuple, cols: dict[str, int]) -> bool:
        num_idx = cols.get("num", 0)
        if num_idx < len(row):
            num = row[num_idx]
            if isinstance(num, (int, float)) and int(num) > 0:
                return True
            if isinstance(num, str) and num.strip().isdigit():
                return True
        return False

    @classmethod
    def _parse_contract_combo(cls, value) -> tuple[str | None, date | None]:
        text = cls._as_text(value)
        if not text:
            return None, None
        date_val = cls._as_date(text)
        num = None
        m = re.search(r"(?:МК|Ф\.|№)\s*([A-Za-zА-Яа-я0-9.\-/]+)", text, flags=re.IGNORECASE)
        if m:
            num = m.group(0).strip()
            if num.upper().startswith("МК"):
                pass
            elif not num.startswith(("Ф.", "№")):
                num = text[:100]
        else:
            num = text[:100]
        return num, date_val

    @classmethod
    def parse_lighting_plan_xlsx(cls, path: Path) -> list[dict]:
        try:
            import openpyxl
        except ImportError as exc:
            raise ValidationError("Для импорта нужен пакет openpyxl. Установите зависимости.") from exc

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        rows: list[dict] = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            year_match = re.search(r"(20\d{2})", sheet_name)
            plan_year = int(year_match.group(1)) if year_match else None
            object_kind = cls.detect_object_kind(sheet_name)

            header_row_idx = None
            header_values = None
            for i, row in enumerate(ws.iter_rows(max_row=20, values_only=True), start=1):
                if any(isinstance(c, str) and "Наименование" in c for c in (row or ())):
                    header_row_idx = i
                    header_values = row
                    break
            if header_row_idx is None or header_values is None:
                continue
            cols = cls._header_map(header_values)
            if "name" not in cols:
                continue

            for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
                if not row or not cls._is_data_row(row, cols):
                    continue
                raw_name = row[cols["name"]] if cols["name"] < len(row) else None
                if raw_name is None:
                    continue
                name = re.sub(r"\s+", " ", str(raw_name).strip())
                if not name or name.casefold().startswith("итого"):
                    continue
                if _JUNK_NAME_RE.match(name):
                    continue

                work_type, address = cls.split_type_and_address(name)
                if not address or len(address) < 3:
                    continue

                deadline = None
                if "deadline" in cols and cols["deadline"] < len(row):
                    deadline = cls._as_text(row[cols["deadline"]])

                contractor = None
                if "contractor" in cols and cols["contractor"] < len(row):
                    contractor = cls._as_text(row[cols["contractor"]])

                contract_number = None
                contract_date = None
                if "contract_combo" in cols and cols["contract_combo"] < len(row):
                    contract_number, contract_date = cls._parse_contract_combo(
                        row[cols["contract_combo"]]
                    )
                if "contract_number" in cols and cols["contract_number"] < len(row):
                    contract_number = cls._as_text(row[cols["contract_number"]]) or contract_number
                if "contract_date" in cols and cols["contract_date"] < len(row):
                    contract_date = cls._as_date(row[cols["contract_date"]]) or contract_date

                # Сумма контракта и НМЦК — разные поля; не подставляем одно в другое
                contract_amount = None
                if "contract_amount" in cols and cols["contract_amount"] < len(row):
                    contract_amount = cls._as_decimal(row[cols["contract_amount"]])
                budget_amount = None
                if "budget" in cols and cols["budget"] < len(row):
                    budget_amount = cls._as_decimal(row[cols["budget"]])

                court_decision = None
                if object_kind == WorkObjectKind.COURT.value:
                    if "court_decision" in cols and cols["court_decision"] < len(row):
                        court_decision = cls._as_text(row[cols["court_decision"]])

                result_text = None
                if "result" in cols and cols["result"] < len(row):
                    result_text = cls._as_text(row[cols["result"]])
                notes = None
                if "notes" in cols and cols["notes"] < len(row):
                    notes = cls._as_text(row[cols["notes"]])

                sip_meters = None
                if "sip" in cols and cols["sip"] < len(row):
                    sip_meters = cls._as_decimal(row[cols["sip"]])
                poles_count = None
                if "poles" in cols and cols["poles"] < len(row):
                    poles_count = cls._as_int(row[cols["poles"]])
                lights_count = None
                if "lights" in cols and cols["lights"] < len(row):
                    lights_count = cls._as_int(row[cols["lights"]])
                shuno_count = None
                if "shuno" in cols and cols["shuno"] < len(row):
                    shuno_count = cls._as_int(row[cols["shuno"]])

                status = WorkObjectStatus.FREE.value
                # Номер контракта в плане — справочное поле, не блокирует создание проекта в Опоре
                if result_text and re.search(r"выполнен|принят", result_text, re.IGNORECASE):
                    status = WorkObjectStatus.COMPLETED.value

                rows.append(
                    {
                        "name": name[:1000],
                        "work_type": work_type,
                        "object_kind": object_kind,
                        "address": address[:1000],
                        "plan_year": plan_year,
                        "work_deadline": deadline,
                        "contract_number": (contract_number or "")[:100] or None,
                        "contract_date": contract_date,
                        "contractor_name": contractor,
                        "contract_amount": contract_amount,
                        "budget_amount": budget_amount,
                        "court_decision_number": (court_decision or "")[:255] or None,
                        "result_text": result_text,
                        "source_sheet": sheet_name.strip()[:100],
                        "notes": notes,
                        "status": status,
                        "sip_meters": sip_meters,
                        "poles_count": poles_count,
                        "lights_count": lights_count,
                        "shuno_count": shuno_count,
                    }
                )
        wb.close()
        return rows

    @classmethod
    def import_from_lighting_plan(cls, path: Path | str, user_id: uuid.UUID) -> ImportResult:
        path = Path(path)
        if not path.is_file():
            raise ValidationError("Файл импорта не найден.")

        parsed = cls.parse_lighting_plan_xlsx(path)
        result = ImportResult(total=len(parsed))
        if not parsed:
            raise ValidationError("В файле не найдено ни одного объекта.")

        # Уникальность: адрес + год + лист (один адрес может быть в разных разделах)
        by_key: dict[str, dict] = {}
        for item in parsed:
            key = "|".join(
                [
                    (item["address"] or "").casefold(),
                    str(item.get("plan_year") or ""),
                    (item.get("source_sheet") or "").casefold(),
                ]
            )
            by_key[key] = item

        existing = {
            "|".join(
                [
                    (obj.address or obj.name or "").casefold(),
                    str(obj.plan_year or ""),
                    (obj.source_sheet or "").casefold(),
                ]
            ): obj
            for obj in db.session.scalars(
                db.select(WorkObject).where(WorkObject.active_filter())
            ).all()
        }

        for key, data in by_key.items():
            obj = existing.get(key)
            if obj is None:
                obj = WorkObject(
                    name=data["name"],
                    work_type=data["work_type"],
                    object_kind=data["object_kind"],
                    address=data["address"],
                    plan_year=data["plan_year"],
                    work_deadline=data["work_deadline"],
                    contract_number=data["contract_number"],
                    contract_date=data["contract_date"],
                    contractor_name=data["contractor_name"],
                    contract_amount=data["contract_amount"],
                    budget_amount=data["budget_amount"],
                    court_decision_number=data["court_decision_number"],
                    result_text=data["result_text"],
                    source_sheet=data["source_sheet"],
                    notes=data["notes"],
                    status=data["status"],
                    created_by=user_id,
                    updated_by=user_id,
                )
                db.session.add(obj)
                result.created += 1
            else:
                obj.name = data["name"]
                obj.work_type = data["work_type"]
                obj.object_kind = data["object_kind"]
                obj.address = data["address"]
                obj.plan_year = data["plan_year"]
                obj.work_deadline = data["work_deadline"]
                obj.contract_number = data["contract_number"]
                obj.contract_date = data["contract_date"]
                obj.contractor_name = data["contractor_name"]
                obj.contract_amount = data["contract_amount"]
                obj.budget_amount = data["budget_amount"]
                obj.court_decision_number = data["court_decision_number"]
                obj.result_text = data["result_text"]
                obj.source_sheet = data["source_sheet"]
                obj.notes = data["notes"]
                obj.status = data["status"]
                obj.updated_by = user_id
                result.updated += 1
            cls._ensure_chain_for_result(
                obj,
                user_id,
                volumes={
                    "sip_meters": data.get("sip_meters"),
                    "poles_count": data.get("poles_count"),
                    "lights_count": data.get("lights_count"),
                    "shuno_count": data.get("shuno_count"),
                },
            )

        AuditService.log(
            user_id=user_id,
            action=AuditAction.CREATE.value,
            entity_type=EntityType.WORK_OBJECT.value,
            description=(
                f"Импорт объектов из плана освещения: "
                f"+{result.created} / ~{result.updated}"
            ),
            new_values={
                "created": result.created,
                "updated": result.updated,
                "total_rows": result.total,
                "unique": len(by_key),
            },
        )
        db.session.commit()
        return result

    @classmethod
    def wipe_all(cls, user_id: uuid.UUID) -> tuple[int, int]:
        """Мягко удалить свободные/архивные объекты. Занятые в цепочке пропускаются."""
        items = list(
            db.session.scalars(db.select(WorkObject).where(WorkObject.active_filter())).all()
        )
        removed = 0
        skipped = 0
        busy = {
            WorkObjectStatus.IN_PROJECT.value,
            WorkObjectStatus.IN_TENDER.value,
            WorkObjectStatus.IN_CONTRACT.value,
        }
        for obj in items:
            if obj.status in busy:
                skipped += 1
                continue
            obj.soft_delete(user_id)
            removed += 1
        if removed:
            AuditService.log(
                user_id=user_id,
                action=AuditAction.SOFT_DELETE.value,
                entity_type=EntityType.WORK_OBJECT.value,
                description=f"Массовое удаление объектов: {removed} (пропущено {skipped})",
                new_values={"count": removed, "skipped": skipped},
            )
            db.session.commit()
        return removed, skipped

    @classmethod
    def _compose_full_name(cls, work_type: str | None, address: str) -> str:
        """Собрать полное наименование, если его не задали вручную."""
        wt = cls._normalize(work_type) or WORK_TYPE_DEFAULT
        addr = address.strip()
        if not addr:
            return wt[:1000]
        if addr.casefold().startswith(wt.casefold()):
            return addr[:1000]
        return f"{wt} {addr}"[:1000]

    @classmethod
    def create(cls, payload: ObjectPayload, user_id: uuid.UUID) -> WorkObject:
        address = cls._normalize(payload.address) or cls._normalize(payload.name) or ""
        if not address:
            raise ValidationError("Адрес объекта обязателен.")
        full_name = cls._normalize(payload.name) or cls._compose_full_name(payload.work_type, address)
        court = cls._normalize(payload.court_decision_number)
        kind = payload.object_kind or WorkObjectKind.PLANNED.value
        if kind != WorkObjectKind.COURT.value:
            court = None
        comment = cls._normalize(payload.kind_comment)
        obj = WorkObject(
            name=full_name[:1000],
            work_type=cls._normalize(payload.work_type) or WORK_TYPE_DEFAULT,
            object_kind=kind,
            kind_comment=(comment[:500] if kind == WorkObjectKind.OTHER.value and comment else None),
            address=address[:1000],
            plan_year=payload.plan_year,
            work_deadline=cls._normalize(payload.work_deadline),
            contract_number=cls._normalize(payload.contract_number),
            contract_date=payload.contract_date,
            contractor_name=cls._normalize(payload.contractor_name),
            contract_amount=payload.contract_amount,
            budget_amount=payload.budget_amount,
            court_decision_number=court,
            result_text=cls._normalize(payload.result_text),
            source_sheet=cls._normalize(payload.source_sheet),
            notes=cls._normalize(payload.notes),
            status=payload.status or WorkObjectStatus.FREE.value,
            created_by=user_id,
            updated_by=user_id,
        )
        db.session.add(obj)
        db.session.flush()
        try:
            if payload.create_draft_project:
                cls._ensure_draft_project(obj, user_id)
            cls._ensure_project_for_result(obj, user_id)
            AuditService.log(
                user_id=user_id,
                action=AuditAction.CREATE.value,
                entity_type=EntityType.WORK_OBJECT.value,
                entity_id=obj.id,
                description=f"Создан объект {obj.display_address}",
                new_values={"address": obj.address, "status": obj.status},
            )
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise
        return obj

    @classmethod
    def update(cls, obj: WorkObject, payload: ObjectPayload, user_id: uuid.UUID) -> WorkObject:
        address = cls._normalize(payload.address) or cls._normalize(payload.name) or ""
        if not address:
            raise ValidationError("Адрес объекта обязателен.")
        full_name = cls._normalize(payload.name) or cls._compose_full_name(payload.work_type, address)
        old = {"address": obj.address, "status": obj.status, "contract_number": obj.contract_number}
        old_result_text = cls._normalize(obj.result_text)
        court = cls._normalize(payload.court_decision_number)
        kind = payload.object_kind or WorkObjectKind.PLANNED.value
        if kind != WorkObjectKind.COURT.value:
            court = None
        comment = cls._normalize(payload.kind_comment)
        obj.name = full_name[:1000]
        obj.work_type = cls._normalize(payload.work_type) or WORK_TYPE_DEFAULT
        obj.object_kind = kind
        if kind == WorkObjectKind.OTHER.value:
            obj.kind_comment = comment[:500] if comment else None
        elif comment:
            obj.kind_comment = comment[:500]
        obj.address = address[:1000]
        obj.plan_year = payload.plan_year
        obj.work_deadline = cls._normalize(payload.work_deadline)
        obj.contract_number = cls._normalize(payload.contract_number)
        obj.contract_date = payload.contract_date
        obj.contractor_name = cls._normalize(payload.contractor_name)
        obj.contract_amount = payload.contract_amount
        obj.budget_amount = payload.budget_amount
        obj.court_decision_number = court
        obj.result_text = cls._normalize(payload.result_text)
        obj.source_sheet = cls._normalize(payload.source_sheet)
        obj.notes = cls._normalize(payload.notes)
        obj.status = payload.status
        obj.updated_by = user_id
        try:
            # Обычное редактирование (сумма, адрес, комментарий) не должно
            # повторно запускать цепочку Project/Tender/Contract. Она зависит
            # от результата работ и запускается только при его изменении.
            if old_result_text != obj.result_text:
                project = cls._ensure_project_for_result(obj, user_id)
                if project is not None:
                    obj.status = WorkObjectStatus.IN_PROJECT.value
            else:
                # Форма может прислать устаревший select status. Обычное
                # редактирование не имеет права откатывать lifecycle объекта.
                obj.status = old["status"]
            AuditService.log(
                user_id=user_id,
                action=AuditAction.UPDATE.value,
                entity_type=EntityType.WORK_OBJECT.value,
                entity_id=obj.id,
                description=f"Обновлён объект {obj.display_address}",
                old_values=old,
                new_values={"address": obj.address, "status": obj.status},
            )
            db.session.commit()
        except Exception:
            db.session.rollback()
            raise
        return obj

    @classmethod
    def soft_delete(cls, obj: WorkObject, user_id: uuid.UUID) -> None:
        if obj.status not in (WorkObjectStatus.FREE.value, WorkObjectStatus.ARCHIVED.value, WorkObjectStatus.COMPLETED.value):
            # Разрешаем удалять и выполненные / из контракта при ручной очистке через wipe
            if obj.status in (
                WorkObjectStatus.IN_PROJECT.value,
                WorkObjectStatus.IN_TENDER.value,
            ):
                raise ValidationError("Нельзя удалить объект, занятый в проекте или на торгах.")
        obj.soft_delete(user_id)
        AuditService.log(
            user_id=user_id,
            action=AuditAction.SOFT_DELETE.value,
            entity_type=EntityType.WORK_OBJECT.value,
            entity_id=obj.id,
            description=f"Удалён объект {obj.display_address}",
        )
        db.session.commit()
