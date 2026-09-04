"""Заполнение бланка-распоряжения на основе неизменяемого XLSX-шаблона."""

from __future__ import annotations

import re
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.styles import Alignment

TEMPLATE_PATH = Path(__file__).with_name("resources") / "order_template.xlsx"
WORK_ROWS = tuple(range(18, 31)) + tuple(range(36, 53))
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def excel_text(value: str | None) -> str:
    """Обычный пользовательский текст не должен становиться Excel-формулой."""
    text = (value or "").strip()
    return f"'{text}" if text.startswith(("=", "+", "-", "@")) else text


def order_filename(number: str | None) -> str:
    safe_number = re.sub(r"[^0-9A-Za-zА-Яа-яЁё._-]+", "_", (number or "").strip()).strip("._") or "без_номера"
    date_label = datetime.now(ZoneInfo("Europe/Moscow")).strftime("%d.%m.%Y")
    return f"Распоряжение_№{safe_number}_{date_label}.xlsx"


def master_initials(value: str | None) -> str:
    """Иванов Иван Петрович -> Иванов И.П.; пустое значение оставляет поле ручным."""
    parts = re.findall(r"[^\s]+", (value or "").strip())
    if len(parts) < 2:
        return (value or "").strip()
    return f"{parts[0]} " + "".join(f"{part[0]}." for part in parts[1:3] if part)


def _replace_signature(label: str, value: str) -> str:
    return f"{label} {excel_text(value)}________________________" if value else f"{label} ________________________"


def _extend_work_rows(sheet, item_count: int) -> tuple[int, ...]:
    if item_count <= len(WORK_ROWS):
        return WORK_ROWS
    extra = item_count - len(WORK_ROWS)
    footer_row = 53
    sheet.insert_rows(footer_row, extra)
    source_row = footer_row - 1
    for row in range(footer_row, footer_row + extra):
        sheet.row_dimensions[row].height = sheet.row_dimensions[source_row].height
        for col in range(1, 8):
            source = sheet.cell(source_row, col)
            target = sheet.cell(row, col)
            target._style = copy(source._style)
            target.number_format = source.number_format
            target.protection = copy(source.protection)
            target.alignment = copy(source.alignment)
    sheet.print_area = f"A1:G{59 + extra}"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    return tuple(range(18, 31)) + tuple(range(36, 53 + extra))


def build_order_workbook(plan: dict, fields: dict[str, str]) -> bytes:
    if not TEMPLATE_PATH.is_file():
        raise ValueError("Шаблон бланка-распоряжения не найден.")
    workbook = load_workbook(TEMPLATE_PATH)
    sheet = workbook["табель"]
    number = excel_text(fields.get("order_number"))
    sheet["D4"] = f"Бланк-распоряжение №{number}" if number else "Бланк-распоряжение №_____________"
    for cell, key in (("D7", "producer"), ("F7", "crew_lead"), ("F9", "crew_members"), ("F10", "lift_responsible")):
        if fields.get(key):
            sheet[cell] = excel_text(fields[key])
    if fields.get("crew_count"):
        sheet["D9"] = excel_text(fields["crew_count"])

    sheet["A54"] = _replace_signature("5. Бланк-распоряжение выдал", fields.get("issuer", ""))
    sheet["A56"] = _replace_signature("7. Целевой-инструктаж провел", fields.get("briefing_conductor", "")) + "                                         8. Целевой инструктаж получил ________________"

    work_rows = _extend_work_rows(sheet, len(plan["items"]))
    for row in work_rows:
        for col in range(1, 8):
            sheet.cell(row, col).value = None
    for index, item in enumerate(plan["items"]):
        row = work_rows[index]
        sheet.cell(row, 1).value = index + 1
        # Колонка «номера» в утверждённом бланке — видимый номер работы,
        # а не ПП. ПП остаётся рядом с адресом и не теряется.
        sheet.cell(row, 2).value = excel_text(item.get("number") or "")
        address = item.get("address") or ""
        pp = item.get("pp") or ""
        pp_label = pp if pp.casefold().startswith("пп") else f"ПП {pp}"
        sheet.cell(row, 3).value = excel_text(f"{pp_label} — {address}" if pp else address)
        sheet.cell(row, 4).value = excel_text(item.get("description") or "")
        for col in range(1, 8):
            cell = sheet.cell(row, col)
            cell.alignment = Alignment(horizontal="left" if col in {3, 4} else "center", vertical="top", wrap_text=True)
        text_size = max(len(str(item.get("address") or "")), len(str(item.get("description") or "")))
        sheet.row_dimensions[row].height = max(sheet.row_dimensions[row].height or 15, min(72, 18 + 12 * ((text_size + 28) // 29)))
    stream = BytesIO()
    workbook.save(stream)
    return stream.getvalue()
