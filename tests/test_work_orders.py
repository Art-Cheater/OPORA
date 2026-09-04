"""Рабочее место мастера: список работ, nearby, план, RBAC."""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook

from app.extensions import db
from app.models.defects.defect import Defect
from app.models.defects.defect_category import DefectCategory
from app.models.defects.defect_status import DefectStatus
from app.models.enums import Priority
from app.models.requests.request import Request
from app.models.requests.request_status import RequestStatus
from app.models.waybills.waybill_stop import WaybillStop
from app.models.waybills.waybill import Waybill
from app.models.work_plans.work_plan import WorkPlan
from app.modules.requests.address_format import normalize_address
from app.modules.requests.repositories import RequestRepository
from app.modules.work_orders.order_service import WORK_ROWS, build_order_workbook


def _login(client, email: str, password: str = "pass12345"):
    client.post("/auth/logout", follow_redirects=True)
    client.post(
        "/auth/login",
        data={"email": email, "password": password, "submit": "Войти"},
        follow_redirects=True,
    )


def _seed_work(app, *, suffix: str):
    with app.app_context():
        journal_id = RequestRepository.get_default_journal().id
        st_new = db.session.scalar(db.select(RequestStatus).where(RequestStatus.code == "new"))
        d_open = db.session.scalar(db.select(DefectStatus).where(DefectStatus.code == "open"))
        category = db.session.scalar(db.select(DefectCategory).where(DefectCategory.code == "lighting"))
        req = Request(
            number=f"26-{suffix}1",
            title="Рядом с дефектом",
            description="Не горит светильник",
            address="ул. Ленина, 27",
            normalized_address=normalize_address("ул. Ленина, 27"),
            street="ул. Ленина",
            district="Ленинский",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=journal_id,
            latitude=Decimal("58.6036"),
            longitude=Decimal("49.6681"),
        )
        extra = Request(
            number=f"26-{suffix}2",
            title="Дальше по улице",
            description="Кабель",
            address="ул. Ленина, 31",
            normalized_address=normalize_address("ул. Ленина, 31"),
            street="ул. Ленина",
            district="Ленинский",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=journal_id,
            latitude=Decimal("58.6038"),
            longitude=Decimal("49.6684"),
        )
        defect = Defect(
            number=f"DF-26-{suffix}",
            description="Не работает светильник",
            address="ул. Ленина, 25",
            normalized_address=normalize_address("ул. Ленина, 25"),
            street="ул. Ленина",
            district="Ленинский",
            status_id=d_open.id,
            category_id=category.id,
            latitude=Decimal("58.6035"),
            longitude=Decimal("49.6680"),
        )
        db.session.add_all([req, extra, defect])
        db.session.commit()
        return str(req.id), str(extra.id), str(defect.id), req.status_id, defect.status_id


def test_work_orders_access(client):
    page = client.get("/work-orders/", follow_redirects=False)
    assert page.status_code in (302, 401)
    _login(client, "master@test.local")
    page = client.get("/work-orders/")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "Работа с заявками" in html or "Работа по заявкам" in html
    assert 'id="workOrderRoot"' in html
    assert "Доступные работы" in html
    assert "Мой план работ" in html
    assert "Номер ПП" in html
    assert "Тип работы" in html
    assert "Из деревень" in html
    assert "Мои планы" in html
    assert "Создать план" in html
    assert "Создать путевой лист" not in html
    assert "js/work-orders.js" in html
    assert 'id="opsMap"' in html
    assert "js/ops-map.js" in html
    assert "vendor/leaflet/leaflet.js" in html
    assert "/work-orders/plans/new" in html
    _login(client, "executor@test.local")
    assert client.get("/work-orders/").status_code == 200
    denied = client.post("/work-orders/plan/add", json={"entity_type": "defect", "entity_id": "00000000-0000-0000-0000-000000000001"})
    assert denied.status_code == 403


def test_order_blank_is_filled_from_the_real_template():
    content = build_order_workbook(
        {
            "items": [
                {"pp": "ПП-12", "address": "ул. Лепсе, 12", "description": "Не горит светильник"},
                {"pp": "", "address": "ул. Лепсе, 15", "description": "Обрыв провода"},
            ]
        },
        {
            "order_number": "7",
            "producer": "Иванов И.И.",
            "crew_lead": "Петров П.П.",
            "crew_members": "Сидоров С.С.",
            "lift_responsible": "Кузнецов К.К.",
        },
    )
    sheet = load_workbook(BytesIO(content), data_only=False)["табель"]
    assert sheet["D4"].value == "Бланк-распоряжение №7"
    assert sheet["D7"].value == "Иванов И.И."
    assert sheet.cell(WORK_ROWS[0], 2).value == "ПП-12"
    assert sheet.cell(WORK_ROWS[1], 3).value == "ул. Лепсе, 15"


def test_work_desk_queue_card_and_complete(client, app):
    _login(client, "master@test.local")
    request_id, extra_id, defect_id, _, _ = _seed_work(app, suffix="91")
    queue = client.get("/work-orders/queue.json").get_json()
    ids = {row["id"] for row in queue["items"]}
    assert request_id in ids
    assert extra_id in ids
    assert defect_id in ids
    assert all(item["entity_type"] in {"request", "defect"} for item in queue["items"])
    requests_only = client.get("/work-orders/queue.json?journal=requests").get_json()
    req_ids = {item["id"] for item in requests_only["items"]}
    assert request_id in req_ids
    assert extra_id in req_ids
    assert defect_id not in req_ids
    defects_only = client.get("/work-orders/queue.json?journal=defects").get_json()
    def_ids = {item["id"] for item in defects_only["items"]}
    assert defect_id in def_ids
    assert request_id not in def_ids
    assert extra_id not in def_ids
    assert all(item["entity_type"] == "defect" for item in defects_only["items"])
    row = next(item for item in queue["items"] if item["id"] == request_id)
    assert row["number"].startswith("26-")
    assert row["address"]
    assert "status" in row
    assert row["can_complete"] is True
    found = client.get(f"/work-orders/queue.json?q=Ленина, 27").get_json()
    found_ids = {item["id"] for item in found["items"]}
    assert request_id in found_ids
    assert extra_id not in found_ids
    card = client.get(f"/work-orders/requests/{request_id}.json").get_json()
    assert card["id"] == request_id
    assert "photos" in card
    assert "history" in card
    assert "description" in card
    assert card["pp"] == "" or card["pp"] is not None
    completed = client.post(f"/work-orders/requests/{request_id}/complete", json={})
    assert completed.status_code == 200, completed.get_data(as_text=True)
    body = completed.get_json()
    assert body["ok"] is True
    assert body["item"]["status_code"] == "completed"
    assert body["card"]["can_complete"] is False
    with app.app_context():
        req = db.session.get(Request, request_id)
        assert db.session.get(RequestStatus, req.status_id).code == "completed"
    done = client.get("/work-orders/queue.json?preset=completed").get_json()
    assert request_id in {item["id"] for item in done["items"]}
    fresh = client.get("/work-orders/queue.json?preset=new").get_json()
    assert request_id not in {item["id"] for item in fresh["items"]}


def test_work_orders_map_colors_and_types(admin_client, app):
    request_id, extra_id, defect_id, _, _ = _seed_work(app, suffix="81")
    initial_ids = {point["id"] for point in admin_client.get("/work-orders/map.json").get_json()["points"]}
    assert {request_id, extra_id, defect_id}.issubset(initial_ids)
    assert admin_client.post("/work-orders/plan/add", json={"entity_type": "request", "entity_id": request_id}).status_code == 200
    assert admin_client.post("/work-orders/plan/add", json={"entity_type": "defect", "entity_id": defect_id}).status_code == 200
    payload = admin_client.get("/work-orders/map.json").get_json()
    points = payload["points"]
    by_id = {p["id"]: p for p in points}
    assert by_id[request_id]["type"] == "request"
    assert by_id[request_id]["color"] == "blue"
    assert by_id[defect_id]["type"] == "defect"
    assert by_id[defect_id]["color"] == "red"


def test_work_orders_plan_nearby_reorder_route(client, app):
    _login(client, "master@test.local")
    request_id, extra_id, defect_id, req_status, def_status = _seed_work(app, suffix="82")
    with app.app_context():
        db.session.get(Request, request_id).pp = "ПП-82"
        db.session.get(Request, extra_id).pp = "ПП-82"
        db.session.get(Defect, defect_id).pp = "ПП-82"
        db.session.commit()
    first = client.post(
        "/work-orders/plan/add",
        json={"entity_type": "defect", "entity_id": defect_id},
    )
    assert first.status_code == 200, first.get_data(as_text=True)
    body = first.get_json()
    assert body["ok"] is True
    assert len(body["plan"]["stops"]) == 1
    nearby_ids = {(h["entity_type"], h["entity_id"]) for h in body["nearby"]["hits"]}
    assert ("request", request_id) in nearby_ids
    assert ("request", extra_id) in nearby_ids
    second = client.post(
        "/work-orders/plan/add",
        json={"entity_type": "request", "entity_id": request_id},
    )
    assert second.status_code == 200
    plan = second.get_json()["plan"]
    assert len(plan["stops"]) == 2
    types = {s["entity_type"] for s in plan["stops"]}
    assert types == {"request", "defect"}
    dup = client.post(
        "/work-orders/plan/add",
        json={"entity_type": "defect", "entity_id": defect_id},
    )
    assert dup.status_code == 400
    stop_ids = [s["id"] for s in plan["stops"]]
    reversed_ids = list(reversed(stop_ids))
    reorder = client.post("/work-orders/plan/reorder", json={"stop_ids": reversed_ids})
    assert reorder.status_code == 200
    ordered = [s["id"] for s in reorder.get_json()["plan"]["stops"]]
    assert ordered == reversed_ids
    with app.app_context():
        selection = db.session.get(Waybill, plan["id"])
        for stop in selection.stops:
            stop.latitude = None
            stop.longitude = None
        db.session.commit()
    route = client.get("/work-orders/route.json").get_json()
    assert len(route["points"]) == 2
    assert route["missing"] == 0
    assert [p["order"] for p in route["points"]] == [1, 2]
    saved = client.post("/work-orders/plan/save", json={})
    assert saved.status_code == 200
    saved_body = saved.get_json()
    assert saved_body["plan"]["number"].startswith("ПР-")
    assert saved_body["plan"]["status"] == "in_progress"
    assert saved_body["redirect"].endswith(saved_body["plan"]["id"])
    already_active = client.post(
        "/work-orders/plan/add",
        json={"entity_type": "request", "entity_id": request_id},
    )
    assert already_active.status_code == 400
    assert "другой активный план" in already_active.get_json()["message"]
    items = client.get("/work-orders/items.json").get_json()["items"]
    by_type = {row["type"] for row in items}
    assert "request" in by_type
    assert "defect" in by_type
    only_def = client.get("/work-orders/items.json?kind=defect").get_json()["items"]
    assert all(row["type"] == "defect" for row in only_def)
    only_req = client.get("/work-orders/items.json?kind=request").get_json()["items"]
    assert all(row["type"] == "request" for row in only_req)
    with app.app_context():
        req = db.session.get(Request, request_id)
        defect = db.session.get(Defect, defect_id)
        assert db.session.get(RequestStatus, req.status_id).code == "in_progress"
        assert db.session.get(DefectStatus, defect.status_id).code == "in_progress"
        mapper_names = {mapper.class_.__name__ for mapper in db.Model.registry.mappers}
        assert "RequestDefect" not in mapper_names
        created_plan = db.session.get(WorkPlan, saved_body["plan"]["id"])
        assert created_plan is not None and created_plan.deleted_at is None
        legacy = db.session.scalar(db.select(Waybill).where(Waybill.master_id == created_plan.master_id))
        assert legacy is not None and legacy.deleted_at is not None


def test_dispatcher_cannot_edit_work_plan(client, app):
    _login(client, "dispatcher@test.local")
    assert client.get("/work-orders/").status_code == 200
    _, _, defect_id, _, _ = _seed_work(app, suffix="83")
    added = client.post(
        "/work-orders/plan/add",
        json={"entity_type": "defect", "entity_id": defect_id},
    )
    assert added.status_code == 403
    assert client.post("/work-orders/plan/complete", json={}).status_code == 403
    assert client.get("/work-orders/plans/new").status_code == 403
    assert client.post("/work-orders/plans/", json={"items": []}).status_code == 403


def test_work_plans_journals_related_complete_and_auto_close(client, app):
    from app.models.work_plans.work_plan import WorkPlan
    from app.modules.requests.journals import JOURNAL_OKTYABRSKY_VILLAGES
    from app.modules.requests.repositories import RequestRepository

    _login(client, "master@test.local")
    with app.app_context():
        main = RequestRepository.get_default_journal()
        villages = RequestRepository.get_journal_by_code(JOURNAL_OKTYABRSKY_VILLAGES)
        st_new = db.session.scalar(db.select(RequestStatus).where(RequestStatus.code == "new"))
        d_open = db.session.scalar(db.select(DefectStatus).where(DefectStatus.code == "open"))
        category = db.session.scalar(db.select(DefectCategory).where(DefectCategory.code == "lighting"))
        first = Request(
            number="25-501",
            title="Опора",
            description="Не горит светильник",
            address="ул. Ленина, 10",
            street="ул. Ленина",
            district="Ленинский",
            pp="69",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=main.id,
        )
        same_pp = Request(
            number="25-512",
            title="Кабель",
            description="Обрыв",
            address="ул. Ленина, 12",
            street="ул. Ленина",
            district="Ленинский",
            pp="ПП 69",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=main.id,
        )
        other_pp = Request(
            number="25-530",
            title="Другой ПП",
            description="Другой фидер",
            address="ул. Мира, 1",
            street="ул. Мира",
            district="Ленинский",
            pp="70",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=main.id,
        )
        village = Request(
            number="25-601",
            title="Деревня",
            description="Деревенская заявка",
            address="д. Широковцы, 2",
            street="Центральная",
            district="Октябрьский",
            pp="12",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=villages.id,
        )
        defect = Defect(
            number="DF-26-14",
            description="Не работает светильник",
            address="ул. Ленина, 8",
            street="ул. Ленина",
            district="Ленинский",
            pp="69",
            status_id=d_open.id,
            category_id=category.id,
        )
        db.session.add_all([first, same_pp, other_pp, village, defect])
        db.session.commit()
        first_id, same_id, other_id, village_id, defect_id = (
            str(first.id),
            str(same_pp.id),
            str(other_pp.id),
            str(village.id),
            str(defect.id),
        )

    html = client.get("/work-orders/").get_data(as_text=True)
    assert "leaflet" in html.lower()
    assert 'id="opsMap"' in html
    assert "Доступные работы" in html

    defects_only = client.get("/work-orders/queue.json?journal=defects").get_json()["items"]
    assert {row["id"] for row in defects_only} == {defect_id}
    assert all(row["number"].startswith("DF-") for row in defects_only)

    village_only = client.get("/work-orders/queue.json?journal=oktyabrsky_villages").get_json()["items"]
    assert {row["id"] for row in village_only} == {village_id}

    with app.app_context():
        from sqlalchemy import func

        from app.models.work_plans.work_plan import WorkPlan

        plans_before = db.session.scalar(
            db.select(func.count()).select_from(WorkPlan).where(WorkPlan.active_filter())
        ) or 0

    create_page = client.get("/work-orders/plans/new")
    assert create_page.status_code == 200, create_page.get_data(as_text=True)
    create_html = create_page.get_data(as_text=True)
    assert "Создание плана работ" in create_html
    assert "Сохранить план" in create_html
    assert "Номер ПП" in create_html
    assert "Из деревень" in create_html
    assert "Доступные работы" in create_html
    assert "черновик" not in create_html.lower()
    with app.app_context():
        from sqlalchemy import func

        from app.models.work_plans.work_plan import WorkPlan

        plans_after_open = db.session.scalar(
            db.select(func.count()).select_from(WorkPlan).where(WorkPlan.active_filter())
        ) or 0
        assert plans_after_open == plans_before

    related = client.get(
        f"/work-orders/related.json?entity_type=request&entity_id={first_id}&skip_request={first_id}"
    ).get_json()
    related_ids = {row["entity_id"] for row in related["by_pp"]}
    assert same_id in related_ids
    assert defect_id in related_ids
    assert first_id not in related_ids
    assert other_id not in related_ids

    created = client.post(
        "/work-orders/plans/",
        json={
            "items": [
                {"entity_type": "request", "entity_id": first_id},
                {"entity_type": "request", "entity_id": same_id},
                {"entity_type": "defect", "entity_id": defect_id},
            ]
        },
    )
    assert created.status_code == 200, created.get_data(as_text=True)
    body = created.get_json()
    saved_plan = body["plan"]
    plan_id = saved_plan["id"]
    assert saved_plan["number"].startswith("ПР-")
    assert saved_plan["status"] == "in_progress"
    assert "draft" not in saved_plan["status"]
    assert body["redirect"].endswith(f"/work-orders/plans/{plan_id}")
    page = client.get(body["redirect"])
    assert page.status_code == 200
    page_html = page.get_data(as_text=True)
    assert saved_plan["number"] in page_html
    assert "В работе" in page_html
    assert "Добавить работы" in page_html
    assert f"/work-orders/plans/{plan_id}/available.json" in page_html
    with app.app_context():
        for entity_id, model in ((first_id, Request), (same_id, Request)):
            row = db.session.get(model, entity_id)
            assert db.session.get(RequestStatus, row.status_id).code == "in_progress"
        defect_row = db.session.get(Defect, defect_id)
        assert db.session.get(DefectStatus, defect_row.status_id).code == "in_progress"

    mine = client.get("/work-orders/plans.json").get_json()["plans"]
    mine_row = next(row for row in mine if row["id"] == plan_id)
    assert mine_row["status"] == "in_progress"
    assert mine_row["remaining"] == 3
    assert mine_row["total"] == 3
    assert "черновик" not in mine_row["status_label"].lower()

    available = client.get(f"/work-orders/plans/{plan_id}/available.json?kind=request").get_json()["items"]
    available_ids = {row["id"] for row in available}
    assert other_id in available_ids
    assert first_id not in available_ids
    assert same_id not in available_ids
    added_later = client.post(
        f"/work-orders/plans/{plan_id}/items",
        json={"entity_type": "request", "entity_id": other_id},
    )
    assert added_later.status_code == 200, added_later.get_data(as_text=True)
    added_plan = added_later.get_json()["plan"]
    assert added_plan["status"] == "in_progress"
    assert len(added_plan["items"]) == 4
    assert "by_district" in added_later.get_json()["related"]
    with app.app_context():
        added_request = db.session.get(Request, other_id)
        assert db.session.get(RequestStatus, added_request.status_id).code == "in_progress"

    plans_page = client.get("/work-orders/plans/")
    assert plans_page.status_code == 200
    plans_html = plans_page.get_data(as_text=True)
    assert saved_plan["number"] in plans_html
    assert "В работе" in plans_html

    opened = client.get(f"/work-orders/plans/{plan_id}.json").get_json()
    assert opened["number"] == saved_plan["number"]
    assert len(opened["items"]) == 4
    by_number = {item["number"]: item for item in opened["items"]}
    complete_id = by_number["25-501"]["id"]
    exclude_id = by_number["25-512"]["id"]
    last_id = by_number["DF-26-14"]["id"]
    added_id = by_number["25-530"]["id"]

    done = client.post(
        f"/work-orders/plans/{plan_id}/items/{complete_id}/complete",
        data={"comment": "Заменил светильник"},
    )
    assert done.status_code == 200, done.get_data(as_text=True)
    excluded = client.post(
        f"/work-orders/plans/{plan_id}/items/{exclude_id}/exclude",
        json={"reason": "no_access", "comment": "Калитка закрыта"},
    )
    assert excluded.status_code == 200, excluded.get_data(as_text=True)
    still_open = excluded.get_json()["plan"]
    assert still_open["status"] == "in_progress"
    excluded_later = client.post(
        f"/work-orders/plans/{plan_id}/items/{added_id}/exclude",
        json={"reason": "other", "comment": "Перенесено на следующий выезд"},
    )
    assert excluded_later.status_code == 200, excluded_later.get_data(as_text=True)
    assert excluded_later.get_json()["plan"]["status"] == "in_progress"

    last = client.post(
        f"/work-orders/plans/{plan_id}/items/{last_id}/complete",
        data={"comment": "Устранён дефект"},
    )
    assert last.status_code == 200, last.get_data(as_text=True)
    finished = last.get_json()["plan"]
    assert finished["status"] == "completed"
    assert finished["completed_at"]
    assert finished["done"] == 2
    assert finished["excluded"] == 2
    history = client.get("/work-orders/plans.json").get_json()["plans"]
    row = next(item for item in history if item["id"] == plan_id)
    assert row["status"] == "completed"
    assert row["remaining"] == 0
    assert row["done"] == 2
    history_page = client.get("/work-orders/plans/").get_data(as_text=True)
    assert saved_plan["number"] in history_page
    assert "Завершён" in history_page
    with app.app_context():
        plan_row = db.session.get(WorkPlan, plan_id)
        assert plan_row.status == "completed"
        assert db.session.get(RequestStatus, db.session.get(Request, first_id).status_id).code == "completed"
        assert db.session.get(RequestStatus, db.session.get(Request, same_id).status_id).code == "new"
        assert db.session.get(RequestStatus, db.session.get(Request, other_id).status_id).code == "new"
        assert db.session.get(DefectStatus, db.session.get(Defect, defect_id).status_id).code == "fixed"

        from app.models.auth.user import User

        recipient_id = str(
            db.session.scalar(db.select(User.id).where(User.email == "executor@test.local"))
        )

    report = client.post(
        f"/work-orders/plans/{plan_id}/report",
        json={"recipient_id": recipient_id},
    )
    assert report.status_code == 200, report.get_data(as_text=True)
    assert report.get_json()["ok"] is True
    completed_page = client.get(f"/work-orders/plans/{plan_id}").get_data(as_text=True)
    assert 'id="planReportOpen"' in completed_page
    assert 'id="planReportModal"' in completed_page
    with app.app_context():
        from app.core.upload_utils import resolve_storage_path
        from app.models.messenger.messenger_message import MessengerMessage

        message = db.session.get(MessengerMessage, report.get_json()["message_id"])
        assert message is not None
        assert message.mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
        with ZipFile(resolve_storage_path(message.storage_key)) as archive:
            xml = archive.read("word/document.xml").decode("utf-8")
        assert saved_plan["number"] in xml
        assert "Закрытые заявки" in xml
        assert "Мастер QA" in xml


def test_work_orders_filters_pp_district_and_villages(client, app):
    from app.modules.requests.journals import JOURNAL_OKTYABRSKY_VILLAGES
    from app.modules.requests.repositories import RequestRepository

    _login(client, "master@test.local")
    request_id, extra_id, defect_id, _, _ = _seed_work(app, suffix="71")
    with app.app_context():
        villages = RequestRepository.get_journal_by_code(JOURNAL_OKTYABRSKY_VILLAGES)
        st_new = db.session.scalar(db.select(RequestStatus).where(RequestStatus.code == "new"))
        village = Request(
            number="26-719",
            title="Деревня",
            description="Деревенская",
            address="д. Широковцы, 2",
            district="Октябрьский",
            pp="12",
            applicant_name="QA",
            priority=Priority.MEDIUM.value,
            status_id=st_new.id,
            journal_id=villages.id,
        )
        req = db.session.get(Request, request_id)
        req.pp = "69"
        extra = db.session.get(Request, extra_id)
        extra.pp = "70"
        db.session.add(village)
        db.session.commit()
        village_id = str(village.id)

    by_pp = client.get("/work-orders/items.json?pp=69").get_json()["items"]
    assert request_id in {row["id"] for row in by_pp}
    assert extra_id not in {row["id"] for row in by_pp}
    by_district = client.get("/work-orders/items.json?district=Ленинский").get_json()["items"]
    ids = {row["id"] for row in by_district}
    assert request_id in ids
    assert village_id not in ids
    villages_only = client.get("/work-orders/items.json?kind=villages").get_json()["items"]
    assert {row["id"] for row in villages_only} == {village_id}
    defects_only = client.get("/work-orders/items.json?kind=defect").get_json()["items"]
    assert {row["id"] for row in defects_only} == {defect_id}


def test_remove_from_plan_restores_status_only_if_system_changed(client, app):
    _login(client, "master@test.local")
    request_id, extra_id, defect_id, req_status, def_status = _seed_work(app, suffix="72")
    added = client.post("/work-orders/plan/add", json={"entity_type": "defect", "entity_id": defect_id})
    assert added.status_code == 200
    stop_id = added.get_json()["plan"]["stops"][0]["id"]
    with app.app_context():
        defect = db.session.get(Defect, defect_id)
        assert db.session.get(DefectStatus, defect.status_id).code == "in_progress"
    removed = client.post("/work-orders/plan/remove", json={"stop_id": stop_id})
    assert removed.status_code == 200
    assert removed.get_json()["plan"]["stops"] == []
    with app.app_context():
        defect = db.session.get(Defect, defect_id)
        assert defect.status_id == def_status
        assert db.session.get(DefectStatus, defect.status_id).code == "open"

    added_request = client.post("/work-orders/plan/add", json={"entity_type": "request", "entity_id": request_id})
    assert added_request.status_code == 200
    request_stop = added_request.get_json()["plan"]["stops"][0]["id"]
    with app.app_context():
        req = db.session.get(Request, request_id)
        assert db.session.get(RequestStatus, req.status_id).code == "in_progress"
        stop = db.session.get(WaybillStop, request_stop)
        assert stop.previous_status_code == "new"
    removed_request = client.post("/work-orders/plan/remove", json={"stop_id": request_stop})
    assert removed_request.status_code == 200
    with app.app_context():
        req = db.session.get(Request, request_id)
        assert req.status_id == req_status
        assert db.session.get(RequestStatus, req.status_id).code == "new"

    with app.app_context():
        extra = db.session.get(Request, extra_id)
        extra.status_id = db.session.scalar(
            db.select(RequestStatus.id).where(RequestStatus.code == "in_progress")
        )
        db.session.commit()
        extra_status = extra.status_id
    added_extra = client.post("/work-orders/plan/add", json={"entity_type": "request", "entity_id": extra_id})
    assert added_extra.status_code == 200
    extra_stop = added_extra.get_json()["plan"]["stops"][0]["id"]
    client.post("/work-orders/plan/remove", json={"stop_id": extra_stop})
    with app.app_context():
        extra = db.session.get(Request, extra_id)
        assert extra.status_id == extra_status


def test_director_tracking_filters_plans(client, app):
    from app.modules.auth.services import AuthService

    request_id, _, defect_id, _, _ = _seed_work(app, suffix="73")
    _login(client, "master@test.local")
    created = client.post(
        "/work-orders/plans/",
        json={
            "items": [
                {"entity_type": "request", "entity_id": request_id},
                {"entity_type": "defect", "entity_id": defect_id},
            ]
        },
    )
    assert created.status_code == 200
    plan_number = created.get_json()["plan"]["number"]

    with app.app_context():
        AuthService.create_user("director@test.local", "pass12345", "Директор QA", "director")

    _login(client, "director@test.local")
    page = client.get("/work-orders/tracking/?status=in_progress")
    assert page.status_code == 200
    html = page.get_data(as_text=True)
    assert "Отслеживание работы" in html
    assert plan_number in html
    assert "Мастер QA" in html
    assert "В работе" in html
    assert 'name="master_id"' in html
    assert 'name="date_from"' in html

    _login(client, "master@test.local")
    assert client.get("/work-orders/tracking/").status_code == 403



